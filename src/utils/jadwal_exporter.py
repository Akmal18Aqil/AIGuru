"""
Jadwal Exporter - Export schedule to Excel
"""
import os
from typing import List, Dict, Any

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def export_jadwal_to_excel(
    jadwal_data: List[Dict[str, Any]], 
    school_name: str = "Sekolah",
    semester: str = "Ganjil 2024/2025",
    output_path: str = None
) -> str:
    """
    Export jadwal to a formatted Excel file.
    Returns the path to the created file.
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl not installed. Run: pip install openpyxl")
    
    if not output_path:
        output_path = os.path.join(os.getcwd(), f"Jadwal_{school_name.replace(' ', '_')}.xlsx")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jadwal Pelajaran"
    
    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = f"JADWAL PELAJARAN - {school_name.upper()}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center
    
    ws.merge_cells('A2:H2')
    ws['A2'] = f"Semester {semester}"
    ws['A2'].alignment = center
    
    # Headers
    headers = ["Hari", "Jam Ke", "Waktu", "Kelas", "Mata Pelajaran", "Guru"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border
    
    # Data Rows
    row = 5
    current_hari = ""
    hari_colors = {
        "Senin": "E2EFDA",
        "Selasa": "FCE4D6",
        "Rabu": "FFF2CC",
        "Kamis": "DDEBF7",
        "Jumat": "E4DFEC",
        "Sabtu": "F2F2F2"
    }
    
    for entry in jadwal_data:
        hari = entry.get('hari', '')
        fill_color = hari_colors.get(hari, "FFFFFF")
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        values = [
            hari if hari != current_hari else "",  # Only show hari once
            entry.get('jam_ke', ''),
            entry.get('waktu', ''),
            entry.get('kelas', ''),
            entry.get('mata_pelajaran', ''),
            entry.get('guru', '')
        ]
        
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.alignment = center
            cell.border = thin_border
            cell.fill = row_fill
        
        current_hari = hari
        row += 1
    
    # Column widths
    widths = [12, 8, 15, 10, 25, 25]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    wb.save(output_path)
    return output_path


def jadwal_to_dataframe(jadwal_data: List[Dict[str, Any]]):
    """
    Convert jadwal data to pandas DataFrame for Streamlit display.
    """
    try:
        import pandas as pd
        return pd.DataFrame(jadwal_data)
    except ImportError:
        return None
